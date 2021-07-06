import openpyxl
import pymongo
from pymongo import MongoClient
url = "mongodb://localhost:27017/ccs"
cluster = MongoClient(url)
db = cluster["ccs"]
collection = db["ifm_mapper_config_data"]
wb = openpyxl.load_workbook("excelToMongoDB/domain_config_data.xlsx")
sheets = wb.sheetnames
S = len(sheets)


def camelCase(s):
    if(len(s) == 0):
        return
    s1 = ''
    s1 += s[0].lower()
    for i in range(1, len(s)):
        if (s[i] == ' '):
            s1 += s[i + 1].upper()
            i += 1
        elif(s[i - 1] != ' '):
            s1 += s[i]
    return(s1)


count = 0
for s in range(0, S):
    currWorkbook = wb[sheets[s]]
    sheetName = sheets[s]
    R = currWorkbook.max_row
    for r in range(1, R+1):
        count = count + 1
        sourceAttribute = currWorkbook.cell(r, 3).value
        reconAttribute = currWorkbook.cell(r, 3).value
        required = currWorkbook.cell(r, 5).value
        domainId = camelCase(sheets[s])
        interface = {
            "key1": "insrd",
            "key2": domainId,
            "isActive": True,
            "data": {
                    "sourceAttribute": sourceAttribute,
                    "reconAttribute": reconAttribute,
                    "config": {
                        "validation": {
                            "required": required
                        }
                    }
            }
        }
        print(count)
        print(sourceAttribute)
        collection.insert_one(interface)


print("Total objects inserted : ")
print(count)
