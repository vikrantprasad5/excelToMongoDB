import openpyxl
import pymongo
from pymongo import MongoClient
url = "mongodb://localhost:27017/ccs"
cluster = MongoClient(url)
db = cluster["ccs"]
collection = db["ifm_attribute_config_data"]

wb = openpyxl.load_workbook("excelToMongoDB/domain_config_data.xlsx")
sheets = wb.sheetnames
S = len(sheets)


def camelCase(s):
    if(len(s) == 0):
        return
    s1 = ''
    s1 += s[0].lower()
    for i in range(1, len(s)):
        if (s[i] == ' '):
            s1 += s[i + 1].upper()
            i += 1
        elif(s[i - 1] != ' '):
            s1 += s[i]
    return(s1)


for s in range(0, S):
    currWorkbook = wb[sheets[s]]
    sheetName = sheets[s]
    R = currWorkbook.max_row
    for r in range(1, R+1):
        sourceAttribute = camelCase(currWorkbook.cell(r, 2).value)
        reconAttribute = camelCase(currWorkbook.cell(r, 2).value)
        currWorkbook.cell(row=r, column=3, value=sourceAttribute)
        wb.save("excelToMongoDB/camel.xlsx")


print(" document objects inserted successfully! ")
