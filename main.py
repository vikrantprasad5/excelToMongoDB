import openpyxl
import pymongo
from pymongo import MongoClient
url = "mongodb://localhost:27017/testDB"
cluster = MongoClient(url)
db = cluster["test"]
collection = db["testCollection"]

wb = openpyxl.load_workbook("xltoMongo/Report.xlsx")
sheets = wb.sheetnames
S = len(sheets)


def camelCase(s):
    if(len(s) == 0):
        return
    s1 = ''
    s1 += s[0].lower()
    for i in range(1, len(s)):
        if (s[i] == ' '):
            s1 += s[i + 1].upper()
            i += 1
        elif(s[i - 1] != ' '):
            s1 += s[i]
    return(s1)


count = 0
for s in range(0, S):
    currWorkbook = wb[sheets[s]]
    sheetName = sheets[s]
    R = currWorkbook.max_row
    for r in range(1, R+1):
        key2 = camelCase(sheets[s])
        key3 = camelCase(currWorkbook.cell(r, 1).value)
        code = camelCase(currWorkbook.cell(r, 1).value)
        displayName = currWorkbook.cell(r, 2).value
        domainId = camelCase(sheets[s])
        interface = {
            "key1": "insrd",
            "key2": key2,
            "key3": key3,
            "isActive": True,
            "data": {
                    "domainId": domainId,
                    "type": "static",
                    "code": code,
                    "displayName": displayName,
                    "datatype": "String",
                    "config": {
                        "dataSourceUrl": ""
                    }
            }
        }
        collection.insert_one(interface)
        count = count + 1
    print(domainId + " done!")


print("Total " + str(count) + " document objects inserted successfully! ")
